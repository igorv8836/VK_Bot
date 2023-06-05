import datetime
import io
import os
import re

import openpyxl
import requests
import vk_api
import bs4
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id
from vk_api import VkUpload
import PIL.Image as Image

import config
import sqlite3

from Schedule import Schedule


class Bot:
    def __init__(self):
        if not os.path.exists('1.xlsx'):
            self.load_schedule_file()
        self.number_week = divmod((datetime.datetime.now() - datetime.datetime(2023, 2, 6)).days, 7)[0] + 1
        self.key_access = config.TOKEN
        self.vk_session = vk_api.VkApi(token=self.key_access)
        self.longpoll = VkLongPoll(self.vk_session)
        self.conn = sqlite3.connect('users.db')
        self.cur = self.conn.cursor()
        self.schedule_data = []
        self.parse_schedule_file()
        self.schedule = Schedule(self.schedule_data)
        self.teacher_user = dict()
        self.upload = VkUpload(self.vk_session)

    def start(self):
        for event in self.longpoll.listen():
            if event.type == VkEventType.MESSAGE_NEW and event.to_me:
                self.handle_message(event)

    @staticmethod
    def load_schedule_file():
        page = requests.get(config.SCHEDULE_URL)
        t = bs4.BeautifulSoup(page.text, "html.parser")
        res = t.find(string="Институт информационных технологий") \
                  .find_parent("div") \
                  .find_parent("div").findAll('a', {'class': 'uk-link-toggle'})[3:6]
        course_r = r'([1-3]) курс'
        for i in res:
            course = i.find('div', 'uk-link-heading').text.lower().strip()
            course_num = re.match(course_r, course)
            if course_num:
                course_num = course_num.group(1)
                f = open(f'{course_num}.xlsx', "wb")
                link = i.get('href')
                resp = requests.get(link)
                f.write(resp.content)
                f.close()

    def parse_schedule_file(self):
        for c in range(3):
            book = openpyxl.load_workbook(f'{c + 1}.xlsx')
            sheet = book.active
            num_cols = sheet.max_column
            last_group_cell = 0
            for i in range(6, num_cols):
                if last_group_cell >= 4:
                    last_group_cell = -1
                    continue
                column = []
                for j in range(2, 88):
                    v = str(sheet.cell(column=i, row=j).value)
                    v = re.sub(r'\n+', '\n', v)
                    if j == 2 and re.match(r'\w{4}-\d{2}-\d{2}', v):
                        last_group_cell = 0
                    column.append(v)
                if last_group_cell != -1:
                    self.schedule_data.append(column)
                    last_group_cell += 1

    def get_user_group(self, user_id):
        try:
            self.cur.execute("SELECT userid, user_group FROM users WHERE userid=?", (user_id,))
        except Exception:
            return None
        row = self.cur.fetchone()
        if row is not None:
            user_id, user_group = row
            return user_group
        else:
            return None

    def add_user_group(self, user_id, new_group):
        self.cur.execute("SELECT userid FROM users WHERE userid=?", (user_id,))
        existing_record = self.cur.fetchone()
        if existing_record:
            self.cur.execute("UPDATE users SET user_group=? WHERE userid=?", (new_group, user_id))
        else:
            self.cur.execute("INSERT INTO users(userid, user_group) VALUES (?, ?)", (user_id, new_group))
        self.conn.commit()

    def send_message(self, user_id, send_message):
        self.vk_session.method('messages.send',
                               {'user_id': user_id, 'message': send_message, 'random_id': get_random_id()})

    def get_weather(self, count, u_id):
        api_key = '167a4ba5a033a64868eea0253be66cc7'
        city_name = 'Moscow'
        url = f'http://api.openweathermap.org/data/2.5/forecast?q={city_name}&lang=ru&appid={api_key}&units=metric'
        response = requests.get(url)
        weather_str = ''
        json = response.json()
        dates = ["06:00:00", "12:00:00", "18:00:00", "00:00:00"]
        dates_str = ['Ночь', 'Утро', 'День', 'Вечер']
        last = ''
        urls = []
        for data in json['list'][:count]:
            weather = data['weather'][0]['main']
            weather_desc = data['weather'][0]['description']
            temperature = data['main']['temp']
            pressure = data['main']['pressure']
            humidity = data['main']['humidity']
            wind_speed = data['wind']['speed']
            wind_direction = data['wind']['deg']
            icon = data['weather'][0]['icon']
            date = data['dt_txt']

            if date[11:] not in dates or (last == dates_str[0] and json['list'][:count].index(data) != 0):
                continue

            match date[11:13]:
                case '00':
                    date = dates_str[0]
                case '06':
                    date = dates_str[1]
                case '12':
                    date = dates_str[2]
                case '18':
                    date = dates_str[3]
            last = date

            if wind_direction >= 337.5 or wind_direction <= 22.5:
                wind_direction = "северный"
            elif wind_direction >= 22.5 or wind_direction <= 67.5:
                wind_direction = "северо-восточный"
            elif wind_direction >= 67.5 or wind_direction <= 112.5:
                wind_direction = "восточный"
            elif wind_direction >= 112.5 or wind_direction <= 157.5:
                wind_direction = "юго-Восточный"
            elif wind_direction >= 157.5 or wind_direction <= 202.5:
                wind_direction = "южный"
            elif wind_direction >= 202.5 or wind_direction <= 247.5:
                wind_direction = "юго-Западный"
            elif wind_direction >= 247.5 or wind_direction <= 292.5:
                wind_direction = "западный"
            elif wind_direction >= 292.5 or wind_direction <= 337.5:
                wind_direction = "северо-Западный"

            wind_text = ''
            if wind_speed >= 29:
                wind_text = 'ураган'
            elif wind_speed >= 19:
                wind_text = 'шторм'
            elif wind_speed >= 12:
                wind_text = 'сильный'
            elif wind_speed >= 2:
                wind_text = 'слабый'
            else:
                wind_text = 'штиль'

            img_url = f'https://openweathermap.org/img/wn/{icon}.png'
            urls.append(img_url)

            weather_str += f'{date}: {weather_desc}\n' \
                           f'Температура: {temperature} °C\n' \
                           f'Давление: {pressure} мм рт.ст.\n' \
                           f'Влажность: {humidity}%\n' \
                           f'Ветер: {wind_text}, {wind_speed} м/c, {wind_direction}\n\n'
        attachments = []
        images = []
        img = Image.new('RGB', (len(urls) * 50, 50))
        img_width = len(urls) * 50
        img_height = 50
        img = Image.new('RGB', (img_width, img_height))

        for i, url in enumerate(urls):
            response = requests.get(url, stream=True)
            image_bytes = io.BytesIO(response.content)
            image = Image.open(image_bytes)
            img.paste(image, (i * 50, 0))

        # сохраняем объединенное изображение в файл image.png
        img.save("image.png")

        # загружаем изображение в VK и отправляем пользователю
        with open("image.png", "rb") as f:
            img_bytes = io.BytesIO(f.read())
        photo = self.upload.photo_messages(photos=img_bytes)[0]
        attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
        self.vk_session.method('messages.send',
                               {'user_id': u_id,
                                'message': weather_str,
                                'random_id': get_random_id(), 'attachment': ','.join(attachments)})

    def handle_message(self, event):
        u_id = event.user_id
        f_message = event.text.lower()
        group_pattern = r"[а-яА-Я]{4}-\d{2}-\d{2}"
        match = re.match(group_pattern, f_message)
        # ввод группы студента
        if match:
            if f_message.upper() not in self.schedule.str_groups:
                self.send_message(u_id, 'Вы ввели неправильную группу')
                return
            else:
                self.add_user_group(u_id, f_message.upper())
                self.send_message(u_id, 'Я запомнил группу - ' + f_message.upper())
                return

        # наличие группы
        group = self.get_user_group(u_id)
        if not group:
            self.send_message(u_id, 'Я не знаю вашей группы, введите свою группу')
            return

        f_message_sp = f_message.split()

        # смена группы по команде "бот ..."
        if len(f_message_sp) >= 2:
            if f_message_sp[0] == 'бот':
                sp = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота']
                if len(f_message_sp) == 2 and f_message_sp[1].upper() in self.schedule.str_groups:
                    self.add_user_group(u_id, f_message_sp[1].upper())
                    self.send_message(u_id, 'Я запомнил группу - ' + f_message_sp[1].upper())
                elif f_message_sp[1] in sp:
                    if len(f_message_sp) == 3 and f_message_sp[2].upper() in self.schedule.str_groups:
                        self.send_message(u_id, 'Четная неделя:\n' +
                                          self.schedule.groups[
                                              self.schedule.str_groups.index(f_message_sp[2].upper())].week[
                                              sp.index(f_message_sp[1])].str(0))
                        self.send_message(u_id, 'Нечетная неделя:\n' +
                                          self.schedule.groups[
                                              self.schedule.str_groups.index(f_message_sp[2].upper())].week[
                                              sp.index(f_message_sp[1])].str(1))
                    elif len(f_message_sp) == 2 and f_message_sp[1] in sp:
                        self.send_message(u_id, 'Четная неделя:\n' +
                                          self.schedule.groups[self.schedule.str_groups.index(group)].week[
                                              sp.index(f_message_sp[1])].str(0))
                        self.send_message(u_id, 'Нечетная неделя:\n' +
                                          self.schedule.groups[self.schedule.str_groups.index(group)].week[
                                              sp.index(f_message_sp[1])].str(1))
                    else:
                        self.send_message(u_id, 'Я не знаю такой команды')
                else:
                    self.send_message(u_id, 'Я не знаю такой команды')
                return

        match f_message:
            case 'начать':
                self.send_message(u_id, config.START_MESSAGE)
                return
            case 'помощь':
                self.send_message(u_id, config.HELP_MESSAGE)
                return
            case 'бот':
                keyboard = VkKeyboard(one_time=False)
                keyboard.add_button('на сегодня', color=VkKeyboardColor.POSITIVE)
                keyboard.add_button('на завтра', color=VkKeyboardColor.NEGATIVE)
                keyboard.add_line()
                keyboard.add_button('на эту неделю', color=VkKeyboardColor.PRIMARY)
                keyboard.add_button('на следующую неделю', color=VkKeyboardColor.PRIMARY)
                keyboard.add_line()
                keyboard.add_button('какая неделя?', color=VkKeyboardColor.SECONDARY)
                keyboard.add_button('какая группа?', color=VkKeyboardColor.SECONDARY)

                self.vk_session.method('messages.send',
                                       {'user_id': u_id,
                                        'message': 'Выберите период',
                                        'random_id': get_random_id(),
                                        'keyboard': keyboard.get_keyboard()})
                return
            case 'на сегодня':
                if datetime.datetime.now().weekday() == 6:
                    self.send_message(u_id, 'Сегодня выходной')
                else:
                    self.send_message(
                        u_id,
                        'Расписание на сегодня\n' + self.schedule.groups[self.schedule.str_groups.index(group)].week[
                            datetime.datetime.now().weekday()].str(self.number_week % 2))
                return
            case 'на завтра':
                if datetime.datetime.now().weekday() == 5:
                    self.send_message(u_id, 'Завтра выходной')
                elif datetime.datetime.now().weekday() == 6:
                    self.send_message(
                        u_id,
                        'Расписание на завтра\n' + self.schedule.groups[self.schedule.str_groups.index(group)].week[
                            0].str(
                            (self.number_week + 1) % 2))
                else:
                    self.send_message(
                        u_id,
                        'Расписание на завтра\n' + self.schedule.groups[self.schedule.str_groups.index(group)].week[
                            (datetime.datetime.now().weekday() + 1) % 7].str(self.number_week % 2))
                return
            case 'на эту неделю':
                self.send_message(
                    u_id,
                    f'Расписание на {self.number_week} неделю\n' + self.schedule.groups[
                        self.schedule.str_groups.index(group)].str(self.number_week % 2))
                return
            case 'на следующую неделю':
                self.send_message(
                    u_id,
                    f'Расписание на {self.number_week + 1} неделю\n' + self.schedule.groups[
                        self.schedule.str_groups.index(group)].str((self.number_week + 1) % 2))
                return
            case 'какая неделя?':
                self.send_message(u_id, 'Сейчас ' + str(self.number_week) + ' неделя')
                return
            case 'какая группа?':
                self.send_message(u_id, 'Расписание для группы ' + group.upper())
                return
            case 'погода':
                keyboard = VkKeyboard(one_time=False)
                keyboard.add_button('сейчас', color=VkKeyboardColor.PRIMARY)
                keyboard.add_button('сегодня', color=VkKeyboardColor.POSITIVE)
                keyboard.add_button('завтра', color=VkKeyboardColor.POSITIVE)
                keyboard.add_line()
                keyboard.add_button('на 5 дней', color=VkKeyboardColor.POSITIVE)

                self.vk_session.method('messages.send',
                                       {'user_id': u_id,
                                        'message': 'Выберите период',
                                        'random_id': get_random_id(),
                                        'keyboard': keyboard.get_keyboard()})
                return
            case 'сейчас':
                api_key = '167a4ba5a033a64868eea0253be66cc7'
                city_name = 'Moscow'
                url = f'http://api.openweathermap.org/data/2.5/weather?q={city_name}&lang=ru&appid={api_key}&units=metric'
                response = requests.get(url)
                if response.status_code == 200:
                    data = response.json()
                    weather_desc = data['weather'][0]['description']
                    temperature = data['main']['temp']
                    pressure = data['main']['pressure']
                    humidity = data['main']['humidity']
                    wind_speed = data['wind']['speed']
                    wind_direction = data['wind']['deg']
                    icon = data['weather'][0]['icon']

                    if wind_direction >= 337.5 or wind_direction <= 22.5:
                        wind_direction = "северный"
                    elif wind_direction >= 22.5 or wind_direction <= 67.5:
                        wind_direction = "северо-восточный"
                    elif wind_direction >= 67.5 or wind_direction <= 112.5:
                        wind_direction = "восточный"
                    elif wind_direction >= 112.5 or wind_direction <= 157.5:
                        wind_direction = "юго-Восточный"
                    elif wind_direction >= 157.5 or wind_direction <= 202.5:
                        wind_direction = "южный"
                    elif wind_direction >= 202.5 or wind_direction <= 247.5:
                        wind_direction = "юго-Западный"
                    elif wind_direction >= 247.5 or wind_direction <= 292.5:
                        wind_direction = "западный"
                    elif wind_direction >= 292.5 or wind_direction <= 337.5:
                        wind_direction = "северо-Западный"

                    wind_text = ''
                    if wind_speed >= 29:
                        wind_text = 'ураган'
                    elif wind_speed >= 19:
                        wind_text = 'шторм'
                    elif wind_speed >= 12:
                        wind_text = 'сильный'
                    elif wind_speed >= 2:
                        wind_text = 'слабый'
                    else:
                        wind_text = 'штиль'

                    img_url = f'https://openweathermap.org/img/wn/{icon}.png'
                    attachments = []
                    image = requests.get(img_url, stream=True)
                    photo = self.upload.photo_messages(photos=image.raw)[0]
                    attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))

                    weather_str = f'Погода в городе Москва: {weather_desc}\n' \
                                  f'Температура: {temperature} °C\n' \
                                  f'Давление: {pressure} мм рт.ст.\n' \
                                  f'Влажность: {humidity}%\n' \
                                  f'Ветер: {wind_text}, {wind_speed} м/c, {wind_direction}'
                    self.vk_session.method('messages.send',
                                           {'user_id': u_id,
                                            'message': weather_str,
                                            'random_id': get_random_id(), 'attachment': ','.join(attachments)})
                return
            case 'сегодня':
                self.get_weather(8, u_id)
                return
            case 'завтра':
                self.get_weather(16, u_id)
                return
            case 'на 5 дней':

                return

        if len(f_message_sp) >= 2 and f_message_sp[0] == 'найти':
            temp = f_message_sp[1]
            if len(f_message_sp) == 3:
                temp += " " + f_message_sp[2]
            teacher_schedule_req = requests.get(f'https://timetable.mirea.ru/api/teacher/search/{temp}')
            if teacher_schedule_req.status_code == 200:
                teacher_schedule = teacher_schedule_req.json()
                if len(teacher_schedule) == 0:
                    self.send_message(u_id, "Такой преподаватель не найден")
                    return
                elif len(teacher_schedule) > 1:
                    keyboard = VkKeyboard(one_time=True)

                    for i in range(len(teacher_schedule)):
                        keyboard.add_button('Найти ' + teacher_schedule[i]['name'], color=VkKeyboardColor.SECONDARY)
                        if i + 1 != len(teacher_schedule):
                            keyboard.add_line()
                    self.vk_session.method('messages.send',
                                           {'user_id': u_id,
                                            'message': 'Выберите нужного преподавателя',
                                            'random_id': get_random_id(),
                                            'keyboard': keyboard.get_keyboard()})
                    return
                else:
                    lessons = sorted(teacher_schedule[0]['lessons'], key=lambda x: (x['weekday'], x['calls']['num']))
                    lessons_text_1 = '---Нечетная неделя---\n'
                    lessons_text_2 = '---Четная неделя---\n'
                    weeks_1 = []
                    weeks_2 = []
                    day_1 = ''
                    day_2 = ''
                    week = ['понедельник', 'вторник', 'среду', 'четверг', 'пятницу', 'субботу']
                    for i in range(6):
                        sp_1 = [z for z in lessons if z['weekday'] == (i + 1) and z['weeks'][0] == 1]
                        sp_2 = [z for z in lessons if z['weekday'] == (i + 1) and z['weeks'][0] == 2]
                        lessons_text_1 += 'Расписание на ' + week[i] + '\n'
                        day_1 += 'Расписание на ' + week[i] + '\n'
                        lessons_text_2 += 'Расписание на ' + week[i] + '\n'
                        day_2 += 'Расписание на ' + week[i] + '\n'
                        for sp_element in sp_1:
                            lessons_text_1 += \
                                str(sp_element['calls']['num']) \
                                + ') ' + sp_element['discipline']['name'] + ', ' + sp_element['group']['name'] + ', ' + \
                                sp_element['lesson_type']['name'] + ', ' + sp_element['room']['name'] + '\n'
                            day_1 += \
                                str(sp_element['calls']['num']) \
                                + ') ' + sp_element['discipline']['name'] + ', ' + sp_element['group']['name'] + ', ' + \
                                sp_element['lesson_type']['name'] + ', ' + sp_element['room']['name'] + '\n'
                        lessons_text_1 += '\n\n'
                        for sp_element in sp_2:
                            lessons_text_2 += \
                                str(sp_element['calls']['num']) \
                                + ') ' + sp_element['discipline']['name'] + ', ' + sp_element['group']['name'] + ', ' + \
                                sp_element['lesson_type']['name'] + ', ' + sp_element['room']['name'] + '\n'
                            day_2 += \
                                str(sp_element['calls']['num']) \
                                + ') ' + sp_element['discipline']['name'] + ', ' + sp_element['group']['name'] + ', ' + \
                                sp_element['lesson_type']['name'] + ', ' + sp_element['room']['name'] + '\n'
                        lessons_text_2 += '\n\n'
                        day_1 += '\n\n'
                        day_2 += '\n\n'
                        weeks_1.append(day_1)
                        weeks_2.append(day_2)
                        day_1 = ''
                        day_2 = ''
                    self.teacher_user[u_id] = [weeks_1, weeks_2, lessons_text_1, lessons_text_2]
                    keyboard = VkKeyboard(one_time=False)
                    keyboard.add_button('на сегодня(п)', color=VkKeyboardColor.POSITIVE)
                    keyboard.add_button('на завтра(п)', color=VkKeyboardColor.NEGATIVE)
                    keyboard.add_line()
                    keyboard.add_button('на эту неделю(п)', color=VkKeyboardColor.PRIMARY)
                    keyboard.add_button('на следующую неделю(п)', color=VkKeyboardColor.PRIMARY)

                    self.vk_session.method('messages.send',
                                           {'user_id': u_id,
                                            'message': f_message,
                                            'random_id': get_random_id(),
                                            'keyboard': keyboard.get_keyboard()})
                    return

        match f_message:
            case 'на сегодня(п)':
                if datetime.datetime.now().weekday() == 6:
                    self.send_message(u_id, 'Сегодня выходной')
                else:
                    self.send_message(u_id, self.teacher_user[u_id][(self.number_week + 1) % 2][
                        datetime.datetime.now().weekday()])
                return
            case 'на завтра(п)':
                if datetime.datetime.now().weekday() == 5:
                    self.send_message(u_id, 'Сегодня выходной')
                elif datetime.datetime.now().weekday() == 6:
                    self.send_message(u_id, self.teacher_user[u_id][self.number_week % 2][0])
                else:
                    self.send_message(u_id, self.teacher_user[u_id][(self.number_week + 1) % 2][
                        datetime.datetime.now().weekday() + 1])
                return
            case 'на эту неделю(п)':
                if self.number_week % 2 == 0:
                    self.send_message(u_id, self.teacher_user[u_id][3])
                else:
                    self.send_message(u_id, self.teacher_user[u_id][2])
                return
            case 'на следующую неделю(п)':
                if (self.number_week + 1) % 2 == 0:
                    self.send_message(u_id, self.teacher_user[u_id][3])
                else:
                    self.send_message(u_id, self.teacher_user[u_id][2])
                return
        self.send_message(u_id, "Неизвестная команда")
        return
